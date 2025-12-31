import tkinter as tk
from tkinter import ttk, messagebox
import mysql.connector
from fpdf import FPDF
import win32print
import arabic_reshaper
from bidi.algorithm import get_display
from jdatetime import datetime as jdatetime
import winsound
from openpyxl import Workbook, load_workbook
import os
import json
import pyttsx3
import threading
from collections import defaultdict
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class FoodReservationApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Food Reservation Printer")
        self.root.geometry("1920x1080")
        self.root.configure(bg="#f0f0f0")

        # اتصال به دیتابیس
        try:
            self.db_connection = mysql.connector.connect(
                host="10.0.26.16",
                user="your_user",
                password="your_password",
                database="mysite_db"
            )
            self.cursor = self.db_connection.cursor()
            print("اتصال به دیتابیس موفقیت‌آمیز بود.")
        except mysql.connector.Error as err:
            messagebox.showerror("خطا", f"خطا در اتصال به دیتابیس: {err}")
            return

        # ایجاد تب‌ها
        self.tab_control = ttk.Notebook(root)
        self.tab_main = ttk.Frame(self.tab_control)
        self.tab_settings = ttk.Frame(self.tab_control)
        self.tab_reports = ttk.Frame(self.tab_control)
        self.tab_control.add(self.tab_main, text="صفحه اصلی")
        self.tab_control.add(self.tab_settings, text="تنظیمات")
        self.tab_control.add(self.tab_reports, text="دریافت رزروها")
        self.tab_control.pack(expand=1, fill="both")

        # ایجاد رابط کاربری در تب‌ها
        self.create_main_tab()
        self.create_settings_tab()
        self.create_reports_tab()

        # متغیرهای عمومی
        self.scanned_users = set()
        self.excel_file = None
        self.json_file = None
        self.food_data_file = "food_data.json"
        self.input_type = None  # 'card' یا 'personal'

        # بارگذاری تاریخ شمسی به طور خودکار
        self.load_current_jalali_date()
        self.load_reservations()

    def create_reports_tab(self):
        # ایجاد رابط کاربری در تب گزارش‌گیری
        self.reports_frame = tk.Frame(self.tab_reports, bg="#f0f0f0")
        self.reports_frame.pack(pady=20)

        # عنوان تب
        self.reports_title = tk.Label(self.reports_frame, text="دریافت گزارش رزروها", bg="#f0f0f0",
                                      font=("Far.Homa", 16, "bold"))
        self.reports_title.pack(pady=20)

        # انتخاب تاریخ
        self.report_month_label = tk.Label(self.reports_frame, text="ماه را انتخاب کنید:", bg="#f0f0f0",
                                           font=("Far.Homa", 14))
        self.report_month_label.pack(pady=10)

        self.report_month_combobox = ttk.Combobox(self.reports_frame, values=[
            "فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور",
            "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"
        ], font=("Far.Homa", 14))
        self.report_month_combobox.pack(pady=10)

        self.report_day_label = tk.Label(self.reports_frame, text="روز را انتخاب کنید:", bg="#f0f0f0",
                                         font=("Far.Homa", 14))
        self.report_day_label.pack(pady=10)

        self.report_day_combobox = ttk.Combobox(self.reports_frame, values=[str(i) for i in range(1, 32)],
                                                font=("Far.Homa", 14))
        self.report_day_combobox.pack(pady=10)

        # دکمه دریافت گزارش
        self.generate_report_button = tk.Button(self.reports_frame, text="دریافت گزارش اکسل",
                                                command=self.generate_food_report,
                                                bg="#2196F3", fg="white", font=("Far.Homa", 14))
        self.generate_report_button.pack(pady=20)

        # برچسب وضعیت
        self.report_status_label = tk.Label(self.reports_frame, text="", bg="#f0f0f0",
                                            font=("Far.Homa", 12))
        self.report_status_label.pack(pady=10)

    def generate_food_report(self):
        # تولید گزارش اکسل از رزروهای غذا
        selected_month = self.report_month_combobox.get()
        selected_day = self.report_day_combobox.get()

        if not selected_month or not selected_day:
            messagebox.showerror("خطا", "لطفاً ماه و روز را انتخاب کنید.")
            return

        persian_months = ["فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور",
                          "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"]
        month_number = persian_months.index(selected_month) + 1
        day_number = int(selected_day)

        try:
            # دریافت رزروها از دیتابیس با اطلاعات کامل
            query = """
            SELECT u.user_login, 
                   MAX(CASE WHEN um1.meta_key = 'first_name' THEN um1.meta_value END) as first_name,
                   MAX(CASE WHEN um2.meta_key = 'last_name' THEN um2.meta_value END) as last_name,
                   r.meal_selected
            FROM wp_food_reservations r
            JOIN wp_users u ON r.user_id = u.ID
            LEFT JOIN wp_usermeta um1 ON u.ID = um1.user_id AND um1.meta_key = 'first_name'
            LEFT JOIN wp_usermeta um2 ON u.ID = um2.user_id AND um2.meta_key = 'last_name'
            WHERE r.month = %s AND r.day_of_month = %s
            GROUP BY u.user_login, r.meal_selected
            ORDER BY r.meal_selected, u.user_login
            """
            self.cursor.execute(query, (month_number, day_number))
            reservations = self.cursor.fetchall()

            if not reservations:
                messagebox.showinfo("اطلاعات", "هیچ رزروی برای این تاریخ یافت نشد.")
                return

            # ایجاد فایل اکسل
            report_file = f"گزارش_کامل_رزروهای_{day_number}_{selected_month}.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "لیست رزروها"

            # اضافه کردن هدرها
            headers = ["کد پرسنلی", "نام", "نام خانوادگی", "نوع غذا"]
            ws.append(headers)

            # تنظیم استایل برای هدرها
            bold_font = Font(bold=True)
            for col in range(1, len(headers) + 1):
                ws.cell(row=1, column=col).font = bold_font

            # اضافه کردن داده‌ها و محاسبه تعداد هر غذا
            food_counts = defaultdict(int)
            for row in reservations:
                user_login, first_name, last_name, meal_selected = row
                ws.append([
                    user_login,
                    first_name if first_name else "نامشخص",
                    last_name if last_name else "نامشخص",
                    meal_selected
                ])
                food_counts[meal_selected] += 1

            # ایجاد برگه خلاصه
            summary_sheet = wb.create_sheet(title="جمع بندی غذاها")

            # اضافه کردن هدرهای برگه خلاصه
            summary_sheet.append(["نوع غذا", "تعداد"])
            summary_sheet.cell(row=1, column=1).font = bold_font
            summary_sheet.cell(row=1, column=2).font = bold_font

            # اضافه کردن آمار هر غذا و محاسبه جمع کل
            total = 0
            for food, count in sorted(food_counts.items()):
                summary_sheet.append([food, count])
                total += count

            # اضافه کردن جمع کل
            summary_sheet.append(["جمع کل", total])
            summary_sheet.cell(row=len(food_counts) + 2, column=1).font = bold_font
            summary_sheet.cell(row=len(food_counts) + 2, column=2).font = bold_font

            # تنظیم عرض ستون‌ها
            for sheet in [ws, summary_sheet]:
                for col in sheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    sheet.column_dimensions[column].width = adjusted_width

            # ذخیره فایل
            wb.save(report_file)

            # نمایش پیام موفقیت
            self.report_status_label.config(text=f"گزارش با موفقیت در فایل {report_file} ذخیره شد.", fg="green")

            # باز کردن فایل اکسل
            os.startfile(report_file)

        except Exception as e:
            messagebox.showerror("خطا", f"خطا در تولید گزارش: {e}")
            self.report_status_label.config(text="خطا در تولید گزارش", fg="red")

    def load_current_jalali_date(self):
        current_jalali_date = jdatetime.now()
        current_month = current_jalali_date.month
        current_day = current_jalali_date.day

        self.month_combobox.current(current_month - 1)
        self.day_combobox.current(current_day - 1)
        self.report_month_combobox.current(current_month - 1)
        self.report_day_combobox.current(current_day - 1)

    def create_main_tab(self):
        self.main_frame = tk.Frame(self.tab_main, bg="#f0f0f0")
        self.main_frame.pack(pady=20)

        self.month_label = tk.Label(self.main_frame, text="ماه را انتخاب کنید:", bg="#f0f0f0", font=("Far.Homa", 14))
        self.month_label.pack(pady=10)

        self.month_combobox = ttk.Combobox(self.main_frame, values=[
            "فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور",
            "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"
        ], font=("Far.Homa", 20))
        self.month_combobox.pack(pady=10)

        self.day_label = tk.Label(self.main_frame, text="روز را انتخاب کنید:", bg="#f0f0f0", font=("Far.Homa", 14))
        self.day_label.pack(pady=10)

        self.day_combobox = ttk.Combobox(self.main_frame, values=[str(i) for i in range(1, 32)],
                                         font=("Far.Homa", 14))
        self.day_combobox.pack(pady=10)

        self.load_button = tk.Button(self.main_frame, text="بارگیری رزروها", command=self.request_password_for_loading,
                                     bg="#4CAF50", fg="white", font=("Far.Homa", 14))
        self.load_button.pack(pady=10)

        self.personal_code_label = tk.Label(self.main_frame,
                                            text="کد پرسنلی (6 رقم) یا اسکن کارت (10 رقم با صفر شروع می‌شود):",
                                            bg="#f0f0f0", font=("Far.Homa", 14))
        self.personal_code_label.pack(pady=10)

        self.personal_code_entry = tk.Entry(self.main_frame, font=("Far.Homa", 14))
        self.personal_code_entry.pack(pady=20)
        self.personal_code_entry.bind("<KeyRelease>", self.check_input)
        self.personal_code_entry.focus_set()

        self.message_label = tk.Label(self.main_frame, text="", fg="red", bg="#f0f0f0", font=("Far.Homa", 12))
        self.message_label.pack(pady=5)

    def check_input(self, event):
        input_text = self.personal_code_entry.get().strip()

        # اگر ورودی خالی است، نوع ورودی را ریست می‌کنیم
        if not input_text:
            self.input_type = None
            return

        # بررسی اینکه ورودی فقط عدد باشد
        if not input_text.isdigit():
            self.personal_code_entry.delete(0, tk.END)
            self.message_label.config(text="ورودی نامعتبر! لطفاً فقط عدد وارد کنید.", fg="red")
            self.root.after(2000, lambda: self.message_label.config(text=""))
            return

        # اگر اولین رقم صفر باشد، کد کارت است (10 رقمی)
        if input_text[0] == '0':
            self.input_type = 'card'
            # محدود کردن طول ورودی به 10 رقم
            if len(input_text) > 10:
                self.personal_code_entry.delete(10, tk.END)
                input_text = input_text[:10]
            # اگر طول ورودی به 10 رسید، جستجو انجام می‌شود
            if len(input_text) == 10:
                self.search_user(input_text)
        # اگر اولین رقم غیر صفر باشد، کد پرسنلی است (6 رقمی)
        else:
            self.input_type = 'personal'
            # محدود کردن طول ورودی به 6 رقم
            if len(input_text) > 6:
                self.personal_code_entry.delete(6, tk.END)
                input_text = input_text[:6]
            # اگر طول ورودی به 6 رسید، جستجو انجام می‌شود
            if len(input_text) == 6:
                self.search_user(input_text)

    def search_user(self, code):
        try:
            if not hasattr(self, 'reservations'):
                messagebox.showinfo("اطلاعات", "لطفاً ابتدا رزروها را بارگیری کنید.")
                return

            for reservation in self.reservations:
                user_id, food_name, display_name, user_login = reservation

                # اگر کد کارت باشد با display_name مقایسه می‌کنیم
                if self.input_type == 'card' and str(display_name) == code:
                    self.process_reservation(user_id, food_name, display_name, user_login)
                    return
                # اگر کد پرسنلی باشد با user_login مقایسه می‌کنیم
                elif self.input_type == 'personal' and str(user_login) == code:
                    self.process_reservation(user_id, food_name, display_name, user_login)
                    return

            # اگر کاربر یافت نشد
            messagebox.showinfo("اطلاعات", "کاربر یافت نشد.")
            self.personal_code_entry.delete(0, tk.END)
            self.personal_code_entry.focus_set()
            self.beep(continuous=True)
        except Exception as e:
            messagebox.showerror("خطا", f"خطا در جستجو: {e}")

    def process_reservation(self, user_id, food_name, display_name, user_login):
        if user_id in self.scanned_users:
            first_name, last_name = self.get_user_names(user_id)
            self.show_food_display(first_name, last_name, food_name, already_scanned=True)
            self.beep(continuous=True)
        else:
            self.scanned_users.add(user_id)
            self.save_scanned_users()
            first_name, last_name = self.get_user_names(user_id)
            self.show_food_display(first_name, last_name, food_name)
            self.print_to_pdf_and_printer(user_id, display_name, food_name)
            self.update_excel(user_login, display_name, food_name)

        self.personal_code_entry.delete(0, tk.END)
        self.personal_code_entry.focus_set()

    def create_settings_tab(self):
        self.settings_frame = tk.Frame(self.tab_settings, bg="#f0f0f0")
        self.settings_frame.pack(pady=20)

        self.settings_password_label = tk.Label(self.settings_frame, text="رمز عبور:", bg="#f0f0f0",
                                                font=("Far.Homa", 14))
        self.settings_password_label.pack(pady=10)

        self.settings_password_entry = tk.Entry(self.settings_frame, show="*")
        self.settings_password_entry.pack(pady=10)

        self.settings_login_button = tk.Button(self.settings_frame, text="ورود", command=self.login_to_settings,
                                               bg="#4CAF50", fg="white", font=("Far.Homa", 14))
        self.settings_login_button.pack(pady=10)

        self.reset_date_label = tk.Label(self.settings_frame, text="تاریخ را انتخاب کنید:", bg="#f0f0f0",
                                         font=("Far.Homa", 14))
        self.reset_date_label.pack(pady=10)

        self.reset_month_combobox = ttk.Combobox(self.settings_frame, values=[
            "فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور",
            "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"
        ])
        self.reset_month_combobox.pack(pady=10)

        self.reset_day_combobox = ttk.Combobox(self.settings_frame, values=[str(i) for i in range(1, 32)])
        self.reset_day_combobox.pack(pady=10)

        self.reset_button = tk.Button(self.settings_frame, text="حذف سیو‌های این تاریخ",
                                      command=self.request_password_for_reset, bg="#F44336", fg="white",
                                      font=("XB Niloofar", 14))
        self.reset_button.pack(pady=10)

        self.report_button = tk.Button(self.settings_frame, text="گزارش", command=self.open_report, bg="#2196F3",
                                       fg="white", font=("Far.Homa", 14))
        self.report_button.pack(pady=10)

    def request_password_for_loading(self):
        self.password_window = tk.Toplevel(self.root)
        self.password_window.title("ورود رمز عبور")
        self.password_window.geometry("300x150")
        self.password_window.configure(bg="#f0f0f0")

        self.password_label = tk.Label(self.password_window, text="لطفاً رمز عبور را وارد کنید:", bg="#f0f0f0",
                                       font=("Far.Homa", 12))
        self.password_label.pack(pady=10)

        self.password_entry = tk.Entry(self.password_window, show="*")
        self.password_entry.pack(pady=10)

        self.password_button = tk.Button(self.password_window, text="تأیید", command=self.check_loading_password,
                                         bg="#4CAF50", fg="white", font=("Far.Homa", 12))
        self.password_button.pack(pady=10)

    def check_loading_password(self):
        password = self.password_entry.get()
        if password == "123456":
            self.load_reservations()
            self.password_window.destroy()
        else:
            messagebox.showerror("خطا", "رمز عبور اشتباه است.")

    def request_password_for_reset(self):
        self.reset_password_window = tk.Toplevel(self.root)
        self.reset_password_window.title("ورود رمز عبور")
        self.reset_password_window.geometry("300x150")
        self.reset_password_window.configure(bg="#f0f0f0")

        self.reset_password_label = tk.Label(self.reset_password_window, text="لطفاً رمز عبور را وارد کنید:",
                                             bg="#f0f0f0", font=("Far.Homa", 12))
        self.reset_password_label.pack(pady=10)

        self.reset_password_entry = tk.Entry(self.reset_password_window, show="*")
        self.reset_password_entry.pack(pady=10)

        self.reset_password_button = tk.Button(self.reset_password_window, text="تأیید",
                                               command=self.check_reset_password, bg="#4CAF50", fg="white",
                                               font=("Far.Homa", 12))
        self.reset_password_button.pack(pady=10)

    def check_reset_password(self):
        password = self.reset_password_entry.get()
        if password == "123456":
            self.reset_scanned_users_by_date()
            self.reset_password_window.destroy()
        else:
            messagebox.showerror("خطا", "رمز عبور اشتباه است.")

    def login_to_settings(self):
        password = self.settings_password_entry.get()
        if password == "123456":
            self.settings_password_entry.pack_forget()
            self.settings_password_label.pack_forget()
            self.settings_login_button.pack_forget()
            self.reset_date_label.pack(pady=10)
            self.reset_month_combobox.pack(pady=10)
            self.reset_day_combobox.pack(pady=10)
            self.reset_button.pack(pady=10)
            self.report_button.pack(pady=10)
        else:
            messagebox.showerror("خطا", "رمز عبور اشتباه است.")

    def reset_scanned_users_by_date(self):
        selected_month = self.reset_month_combobox.get()
        selected_day = self.reset_day_combobox.get()

        if not selected_month or not selected_day:
            messagebox.showerror("خطا", "لطفاً ماه و روز را انتخاب کنید.")
            return

        persian_months = ["فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور",
                          "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"]
        month_number = persian_months.index(selected_month) + 1
        day_number = int(selected_day)

        try:
            query = """
            SELECT user_id FROM wp_food_reservations
            WHERE month = %s AND day_of_month = %s
            """
            self.cursor.execute(query, (month_number, day_number))
            users_to_reset = [row[0] for row in self.cursor.fetchall()]

            self.scanned_users = self.scanned_users - set(users_to_reset)
            self.save_scanned_users()
            messagebox.showinfo("اطلاعات", f"سیو‌های تاریخ {selected_day} {selected_month} حذف شدند.")
        except Exception as e:
            messagebox.showerror("خطا", f"خطا در حذف سیو‌ها: {e}")

    def load_scanned_users(self):
        if self.json_file and os.path.exists(self.json_file):
            with open(self.json_file, "r", encoding="utf-8") as f:
                self.scanned_users = set(json.load(f))
        else:
            self.scanned_users = set()

    def save_scanned_users(self):
        with open(self.json_file, "w", encoding="utf-8") as f:
            json.dump(list(self.scanned_users), f, ensure_ascii=False)

    def load_reservations(self):
        threading.Thread(target=self.load_reservations_thread).start()

    def load_reservations_thread(self):
        selected_month = self.month_combobox.get()
        selected_day = self.day_combobox.get()

        if not selected_month or not selected_day:
            messagebox.showerror("خطا", "لطفاً ماه و روز را انتخاب کنید.")
            return

        persian_months = ["فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور",
                          "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"]
        month_number = persian_months.index(selected_month) + 1
        day_number = int(selected_day)

        self.excel_file = f"{day_number} {selected_month}.xlsx"
        self.json_file = f"scanned_users_{day_number}_{month_number}.json"

        self.initialize_excel_file()
        self.load_food_data()

        try:
            query = """
            SELECT r.user_id, r.meal_selected, u.display_name, u.user_login 
            FROM wp_food_reservations r
            JOIN wp_users u ON r.user_id = u.ID
            WHERE r.month = %s AND r.day_of_month = %s
            """
            self.cursor.execute(query, (month_number, day_number))
            self.reservations = self.cursor.fetchall()

            if not self.reservations:
                messagebox.showinfo("اطلاعات", "هیچ رزروی برای این تاریخ یافت نشد.")
            else:
                messagebox.showinfo("اطلاعات", f"{len(self.reservations)} رزرو برای این تاریخ یافت شد.")
                self.load_scanned_users()
                self.personal_code_entry.focus_set()

        except Exception as e:
            messagebox.showerror("خطا", f"خطا در بارگیری رزروها: {e}")

    def load_food_data(self):
        if not os.path.exists(self.food_data_file):
            food_data = {
                "غذاها": [
                    {"id": 1, "name": "غذای 1"},
                    {"id": 2, "name": "غذای 2"},
                    {"id": 3, "name": "غذای 3"},
                ]
            }
            with open(self.food_data_file, "w", encoding="utf-8") as f:
                json.dump(food_data, f, ensure_ascii=False, indent=4)

        with open(self.food_data_file, "r", encoding="utf-8") as f:
            self.food_data = json.load(f)

    def show_food_display(self, first_name, last_name, food_name, already_scanned=False, received_time=None):
        popup = tk.Toplevel(self.root)
        popup.title("اطلاعات غذا")
        popup.geometry("1000x800")
        popup.configure(bg="#f0f0f0")

        name_label = tk.Label(popup, text=f" {first_name} {last_name}", font=("Far.Homa", 72), fg="green",
                              bg="#f0f0f0")
        name_label.pack(pady=40)

        food_label = tk.Label(popup, text=f"{food_name}", font=("Far.Homa", 72), fg="black", bg="#f0f0f0")
        food_label.pack(pady=40)

        if already_scanned:
            message_label = tk.Label(popup, text="شما قبلاً دریافت کرده‌اید.", font=("Far.Homa", 72), fg="red",
                                     bg="#f0f0f0")
            message_label.pack(pady=40)

            if received_time is not None:
                time_label = tk.Label(popup, text=f"شما {received_time} دقیقه قبل غذا را دریافت کرده‌اید.",
                                      font=("Far.Homa", 36), fg="blue", bg="#f0f0f0")
                time_label.pack(pady=20)

        engine = pyttsx3.init()
        engine.setProperty('voice', 'persian')
        engine.say(food_name)
        engine.runAndWait()

        self.root.after(2500, popup.destroy)
        self.personal_code_entry.delete(0, tk.END)
        self.root.after(2500, lambda: self.personal_code_entry.focus_set())

    def print_to_pdf_and_printer(self, user_id, display_name, food_name):
        try:
            first_name, last_name = self.get_user_names(user_id)

            display_name_fixed = get_display(arabic_reshaper.reshape(display_name))
            food_name_fixed = get_display(arabic_reshaper.reshape(food_name))
            first_name_fixed = get_display(arabic_reshaper.reshape(first_name))
            last_name_fixed = get_display(arabic_reshaper.reshape(last_name))

            jalali_date = jdatetime.now().strftime("%Y/%m/%d")

            pdf_file = "food_reservation.pdf"
            pdf = FPDF()
            pdf.add_page()
            pdf.add_font('Persian', '', 'XB Niloofar.ttf', uni=True)
            pdf.set_font('Persian', '', 16)
            pdf.cell(200, 10, txt=f"{jalali_date}", ln=True, align="R")
            pdf.cell(200, 10, txt=f" {first_name_fixed} {last_name_fixed}", ln=True, align="R")
            pdf.cell(200, 10, txt=f" {food_name_fixed}", ln=True, align="R")
            pdf.output(pdf_file)

            self.print_pdf(pdf_file)
            self.show_message("PDF ایجاد و به چاپگر ارسال شد.", "green")
        except Exception as e:
            self.show_message(f"خطا در ایجاد یا چاپ PDF: {e}")

    def print_pdf(self, pdf_file):
        try:
            printer_name = win32print.GetDefaultPrinter()
            hprinter = win32print.OpenPrinter(printer_name)
            printer_handle = win32print.StartDocPrinter(hprinter, 1, ("Food Reservation", None, "RAW"))
            win32print.StartPagePrinter(hprinter)

            with open(pdf_file, "rb") as f:
                pdf_data = f.read()
                win32print.WritePrinter(hprinter, pdf_data)

            win32print.EndPagePrinter(hprinter)
            win32print.EndDocPrinter(hprinter)
            win32print.ClosePrinter(hprinter)

            self.show_message("فایل به چاپگر ارسال شد.", "green")
        except Exception as e:
            self.show_message(f"خطا در چاپ: {e}")

    def initialize_excel_file(self):
        if not os.path.exists(self.excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "گزارش غذا"
            ws.append(["کد پرسنلی", "نام", "غذا", "زمان دریافت"])
            wb.save(self.excel_file)

    def update_excel(self, user_login, display_name, food_name):
        wb = load_workbook(self.excel_file)
        ws = wb.active
        jalali_time = jdatetime.now().strftime("%Y/%m/%d %H:%M:%S")
        ws.append([user_login, display_name, food_name, jalali_time])
        wb.save(self.excel_file)

    def open_report(self):
        if self.excel_file and os.path.exists(self.excel_file):
            os.startfile(self.excel_file)
        else:
            messagebox.showinfo("اطلاعات", "فایل گزارش یافت نشد.")

    def show_message(self, message, color="red"):
        self.message_label.config(text=message, fg=color)
        self.root.after(2000, lambda: self.message_label.config(text=""))

    def beep(self, continuous=False):
        if continuous:
            winsound.Beep(1000, 1000)
        else:
            winsound.Beep(1000, 200)

    def get_user_names(self, user_id):
        query = """
        SELECT 
            MAX(CASE WHEN meta_key = 'first_name' THEN meta_value END) AS first_name,
            MAX(CASE WHEN meta_key = 'last_name' THEN meta_value END) AS last_name
        FROM wp_usermeta 
        WHERE user_id = %s
        """
        self.cursor.execute(query, (user_id,))
        result = self.cursor.fetchone()

        first_name = result[0] if result[0] else "نام نامشخص"
        last_name = result[1] if result[1] else "نام خانوادگی نامشخص"

        return first_name, last_name


if __name__ == "__main__":
    root = tk.Tk()
    app = FoodReservationApp(root)
    root.mainloop()